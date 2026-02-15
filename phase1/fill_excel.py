import openpyxl
import os

file_path = r'c:\Users\huangqi\OneDrive\Desktop\AIMD_HW1\Evaluation Sheet.xlsx'

data_str = """
R1: Task1 / TC1 / Prompt A / ChatGPT5.2 
R2: Task1 / TC1 / Prompt A / Claude Haiku4.5 
R3: Task1 / TC1 / Prompt B / ChatGPT5.2 
R4: Task1 / TC1 / Prompt B / Claude Haiku4.5 
R5: Task1 / TC2 / Prompt A / ChatGPT5.2 
R6: Task1 / TC2 / Prompt A / Claude Haiku4.5 
R7: Task1 / TC2 / Prompt B / ChatGPT5.2 
R8: Task1 / TC2 / Prompt B / Claude Haiku4.5 
R9: Task2 / TC3 / Prompt A / ChatGPT5.2 
R10: Task2 / TC3 / Prompt A / Claude Haiku4.5 
R11: Task2 / TC3 / Prompt B / ChatGPT5.2 
R12: Task2 / TC3 / Prompt B / Claude Haiku4.5 
R13: Task2 / TC4 / Prompt A / ChatGPT5.2 
R14: Task2 / TC4 / Prompt A / Claude Haiku4.5 
R15: Task2 / TC4 / Prompt B / ChatGPT5.2 
R16: Task2 / TC4 / Prompt B / Claude Haiku4.5 
"""

# Parse data
rows_to_add = []
for line in data_str.strip().split('\n'):
    # Format: R1: Task1 / TC1 / Prompt A / ChatGPT5.2
    # Split by ':' first to get ID, then by '/'
    if not line.strip():
        continue
        
    parts = line.split(':')
    run_id = parts[0].strip()
    rest = parts[1].strip()
    
    fields = [f.strip() for f in rest.split('/')]
    # fields should be [Task, TC, Prompt, Model]
    
    row_data = [run_id] + fields
    rows_to_add.append(row_data)

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"Current max row: {ws.max_row}")
    
    # Check if headers match what we expect just to be safe
    # Headers: ['run_id', 'task', 'test_case', 'prompt', 'model', ...]
    # We are filling the first 5 columns
    
    start_row = ws.max_row + 1
    # If the file is completely empty (no headers), max_row is 1 but cell is empty.
    # But we saw headers in previous step.
    
    for i, row_data in enumerate(rows_to_add):
        # row_data: [run_id, task, tc, prompt, model]
        # Write to columns 1 to 5
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=start_row + i, column=col_idx, value=value)
            
    wb.save(file_path)
    print(f"Successfully added {len(rows_to_add)} rows.")
    
except Exception as e:
    print(f"Error: {e}")
